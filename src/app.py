from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

options = Options()
options.add_argument("--incognito")
driver = webdriver.Chrome(options=options)
driver.get('https://www.idealista.pt/comprar-casas/lisboa/')

WebDriverWait(driver, 100).until(
    EC.presence_of_all_elements_located((By.XPATH, "//a[@class='item-link ']"))
)


titles = driver.find_elements(By.XPATH, "//a[@class='item-link ']")
prices = driver.find_elements(By.XPATH, "//span[@class='item-price h2-simulated']")

# for test purpose 
for title in titles: 
    print(title.text)

for price in prices:
    print(price.text)

workbook = openpyxl.Workbook()
workbook.create_sheet('housesIdealista')
sheet_houses = workbook['housesIdealista']

sheet_houses['A1'].value = 'Descrição'
sheet_houses['B1'].value = 'Preço'
sheet_houses['C1'].value = 'Telefone'


try:
    buttons = driver.find_elements(By.XPATH, "//button[@class='icon-phone hidden-contact-phones_link see-phones-btn fake-anchor']")

    for i, button in enumerate(buttons):
        ActionChains(driver).move_to_element(button).click(button).perform()

        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//span[@class='hidden-contact-phones_text']"))
        )

        number = driver.find_elements(By.XPATH, "//span[@class='hidden-contact-phones_text']")

        sheet_houses.cell(row = i + 2, column=1).value = titles[i].text if i < len(titles) else 'N/A'
        sheet_houses.cell(row = i + 2, column=2).value = prices[i].text if i < len(prices) else 'N/A'
        sheet_houses.cell(row = i + 2, column=3).value = number[0].text if number else 'N/A'

except Exception as error:
    print('Something wrong: ', error)

finally: 
    driver.close()
    print('Closing..')

workbook.save('casas_idealista.xlsx')